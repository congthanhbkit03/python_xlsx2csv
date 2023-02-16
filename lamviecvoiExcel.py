import openpyxl

def chuyen_file_upload_users():
    ten_file = input("Nhập tên file (*.xlsx) - vd: dct30 -> chọn file dct30.xlsx: ")
    try:
        wb = openpyxl.load_workbook(ten_file +'.xlsx')  #cai nay se lay tham so sau
        sheet = wb.active   # sheet dau tien

        danhSach = []
        for row in range(6, sheet.max_row + 1):
            account = {}
            account['student_id'] = (sheet['C' + str(row)].value).lower()
            account['username'] = (sheet['C' + str(row)].value).lower()  # account la username chuyen lower
            account['firstname'] = "{}_{}".format(sheet['AA' + str(row)].value, sheet['D' + str(row)].value)
            account['lastname'] = sheet['E' + str(row)].value
            account['email'] = account['username'] + "@email.com"
            account['password'] = account['username']
            danhSach.append((account))

        print(danhSach)

        # chuyen vao thanh file .csv
        with open('upload_users_' + ten_file + ".csv", 'w', encoding="utf-8") as fw:
            fw.write('username,firstname,lastname,email,password\n')
            for acc in danhSach:
                line = acc['username'] + ","
                line += acc['firstname'] + ","
                line += acc['lastname'] + ","
                line += acc['email'] + ","
                line += acc['password'] + "\n"
                fw.write(line)

        print("Tạo file upload user thành công!")
    except Exception:
        print("File không tìm thấy!")

def chuyen_file_upload_enrollment():
    ten_file = input("Nhập tên file chứa tài khoản tham gia môn học (*.xlsx) - vd: tienganh -> chọn file tienganh.xlsx: ")
    try:
        wb = openpyxl.load_workbook(ten_file +'.xlsx')  #cai nay se lay tham so sau
        sheet = wb.active   # sheet dau tien

        danhSach = []
        for row in range(6, sheet.max_row + 1):
            account = {}
            account['username'] = (sheet['C' + str(row)].value).lower()  # account la username chuyen lower
            # account['firstname'] = "{}_{}".format(sheet['AA' + str(row)].value, sheet['D' + str(row)].value)
            # account['lastname'] = sheet['E' + str(row)].value
            # account['email'] = account['username'] + "@email.com"
            # account['password'] = account['username']
            danhSach.append((account))

        print(danhSach)

        # chuyen vao thanh file .csv
        with open("enroll_upload"+ ten_file + '.csv', 'w', encoding="utf-8") as fw:
            fw.write('account\n')
            for acc in danhSach:
                line = acc['username'] + "\n"
                fw.write(line)

        print("Tạo file upload thành viên tham gia thành công!")
    except Exception:
        print("File không tìm thấy!")
menu = """
Chọn chức năng:
1. Chuyển file danh sách Excel (xlsx) sang .CSV để nhập tài khoản cho hệ thống 
2. Tạo file upload thành viên cho lớp học (enrollments)
3. Thoát
"""

def main():
    print(menu)
    while True:
        chucnang = int(input("Nhập chức năng: "))
        if chucnang == 1:
            chuyen_file_upload_users()
        elif chucnang == 2:
            chuyen_file_upload_enrollment()
        elif chucnang == 3:
            print("Exit")
            break
        else:
            print("Chọn lại chức năng!")

main()
